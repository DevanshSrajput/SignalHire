import csv
import json
import pickle
import re
import subprocess
import sys
import time
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill, numbers
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))

import numpy as np

from config import (
    ARTIFACTS_DIR,
    CANDIDATES_PATH,
    OUTPUT_DIR,
    TOP_K,
    VALIDATE_SCRIPT,
    WEIGHTS,
)
from engine import format_score
from evidence import generate_reasoning


_CID_PATTERN = re.compile(rb'"candidate_id"\s*:\s*"([^"]+)"')


def build_offset_index() -> dict:
    """candidate_id -> byte offset of its line in the JSONL. One ~2s pass;
    afterwards any candidate loads with a seek instead of a file scan."""
    index = {}
    with open(CANDIDATES_PATH, "rb") as f:
        while True:
            pos = f.tell()
            line = f.readline()
            if not line:
                break
            m = _CID_PATTERN.search(line)
            if m:
                index[m.group(1).decode("utf-8")] = pos
    return index


def load_candidates_by_ids(target_ids, offset_index: dict = None) -> dict:
    """With an offset index: direct seeks per id. Without: a single pass
    over the JSONL (never one scan per id)."""
    if offset_index is not None:
        found = {}
        with open(CANDIDATES_PATH, "rb") as f:
            for cid in target_ids:
                pos = offset_index.get(cid)
                if pos is None:
                    continue
                f.seek(pos)
                found[cid] = json.loads(f.readline())
        return found

    remaining = set(target_ids)
    found = {}
    with open(CANDIDATES_PATH, "r", encoding="utf-8") as f:
        for line in f:
            if not remaining:
                break
            line = line.strip()
            if not line:
                continue
            cand = json.loads(line)
            cid = cand.get("candidate_id")
            if cid in remaining:
                found[cid] = cand
                remaining.discard(cid)
    return found


def main():
    t0 = time.time()

    print("Loading artifacts ...")
    embeddings = np.load(str(ARTIFACTS_DIR / "embeddings.npy")).astype(np.float32)
    candidate_ids = np.load(str(ARTIFACTS_DIR / "candidate_ids.npy"), allow_pickle=True)
    jd_embedding = np.load(str(ARTIFACTS_DIR / "jd_embedding.npy")).astype(np.float32)

    with open(ARTIFACTS_DIR / "subscores.pkl", "rb") as f:
        subscores_dict = pickle.load(f)

    n = len(candidate_ids)
    print(f"Loaded {n} candidates, embeddings shape: {embeddings.shape}")

    semantic_sim = embeddings @ jd_embedding

    weight_vector = np.array(
        [WEIGHTS["technical_fit"], WEIGHTS["career_quality"], WEIGHTS["availability_signal"], WEIGHTS["seniority_fit"]],
        dtype=np.float32,
    )

    subscore_matrix = np.zeros((n, 4), dtype=np.float32)
    penalty_multipliers = np.ones(n, dtype=np.float32)

    for i, cid in enumerate(candidate_ids):
        ss = subscores_dict.get(cid, {})
        subscore_matrix[i, 0] = ss.get("technical_fit", 0.0)
        subscore_matrix[i, 1] = ss.get("career_quality", 0.0)
        subscore_matrix[i, 2] = ss.get("availability_signal", 0.0)
        subscore_matrix[i, 3] = ss.get("seniority_fit", 0.0)
        penalty_multipliers[i] = ss.get("penalty_multiplier", 1.0)

    print("Computing composite scores ...")
    base_scores = subscore_matrix @ weight_vector
    scores = penalty_multipliers * (base_scores + WEIGHTS["semantic_similarity"] * semantic_sim)

    k = min(TOP_K, len(scores))
    top_indices = np.argsort(-scores)[:k]
    top_k_pairs = [(float(scores[i]), str(candidate_ids[i])) for i in top_indices]

    # Sort on the emitted (rescaled, rounded) score so the CSV is provably
    # non-increasing at output precision, with candidate_id ascending as the
    # tie-break the validator requires for equal scores.
    top_k_pairs.sort(key=lambda x: (-float(format_score(x[0])), x[1]))

    top_ids = np.array([p[1] for p in top_k_pairs], dtype=object)

    print(f"Top score: {top_k_pairs[0][0]:.4f}, bottom: {top_k_pairs[-1][0]:.4f}")

    print("Loading top candidates for reasoning ...")
    candidates_by_id = load_candidates_by_ids(cid for _, cid in top_k_pairs)
    top_candidates = [candidates_by_id.get(cid) for _, cid in top_k_pairs]

    output_path = OUTPUT_DIR / "submission.csv"
    print(f"Writing {output_path} ...")

    with open(output_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(["candidate_id", "rank", "score", "reasoning"])

        for rank_idx, (score_val, cid) in enumerate(top_k_pairs):
            rank = rank_idx + 1
            cand = top_candidates[rank_idx]
            reasoning = generate_reasoning(cand or {})
            writer.writerow([cid, rank, format_score(score_val), reasoning])

    # ── XLSX export ──────────────────────────────────────────────────────────
    xlsx_path = OUTPUT_DIR / "submission.xlsx"
    print(f"Writing {xlsx_path} ...")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rankings"

    # Header style
    hdr_fill = PatternFill("solid", fgColor="1E1B4B")
    hdr_font = Font(bold=True, color="E0E7FF", size=11)
    headers = ["Rank", "Candidate ID", "Score (/10)", "Reasoning"]
    col_widths = [7, 18, 13, 90]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    # Row styles: alternate light purple / white
    fill_even = PatternFill("solid", fgColor="EEF2FF")
    fill_odd  = PatternFill("solid", fgColor="FFFFFF")
    score_font = Font(bold=True, color="4F46E5")

    # Re-read the CSV we just wrote so we don't duplicate logic
    with open(output_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row_idx, row in enumerate(reader, 2):
            fill = fill_even if row_idx % 2 == 0 else fill_odd
            cells = [
                (row["rank"],         Alignment(horizontal="center")),
                (row["candidate_id"], Alignment(horizontal="left")),
                (row["score"],        Alignment(horizontal="center")),
                (row["reasoning"],    Alignment(horizontal="left", wrap_text=True)),
            ]
            for col, (val, align) in enumerate(cells, 1):
                c = ws.cell(row=row_idx, column=col, value=val)
                c.fill = fill
                c.alignment = align
                if col == 3:
                    c.font = score_font
            ws.row_dimensions[row_idx].height = 40

    ws.freeze_panes = "A2"
    wb.save(xlsx_path)
    # ─────────────────────────────────────────────────────────────────────────

    elapsed = time.time() - t0
    print(f"Ranking complete in {elapsed:.1f}s")

    if VALIDATE_SCRIPT.exists():
        print("Validating submission ...")
        result = subprocess.run(
            [sys.executable, str(VALIDATE_SCRIPT), str(output_path)],
            capture_output=True, text=True,
        )
        print(result.stdout)
        if result.stderr:
            print("STDERR:", result.stderr)
        if result.returncode != 0:
            print("VALIDATION FAILED")
            sys.exit(1)
        print("Submission valid!")
    else:
        print("validate_submission.py not found — skipping in-container validation.")
    print(f"Output (CSV):  {output_path}")
    print(f"Output (XLSX): {xlsx_path}")


if __name__ == "__main__":
    main()
